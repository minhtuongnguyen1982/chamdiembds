import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import io
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from lead_scorer import analyze_and_score_lead

# Page setup
st.set_page_config(
    page_title="AI Lead Scoring & Automation - Bất Động Sản",
    page_icon="🏠",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for Premium Dark UI
st.markdown("""
<style>
    /* Dark Theme Custom Palette */
    .stApp {
        background-color: #0B0F17;
        color: #E2E8F0;
    }
    
    /* Header Container */
    .header-box {
        background: linear-gradient(135deg, #1E293B 0%, #0F172A 100%);
        border: 1px solid #334155;
        border-radius: 16px;
        padding: 24px;
        margin-bottom: 24px;
        box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.4);
    }
    
    /* Metric Cards */
    .metric-card {
        background: linear-gradient(145deg, #1E2640 0%, #111827 100%);
        border: 1px solid #374151;
        border-radius: 12px;
        padding: 18px 20px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.25);
    }
    .metric-title {
        font-size: 0.82rem;
        color: #94A3B8;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        font-weight: 600;
    }
    .metric-value {
        font-size: 1.8rem;
        font-weight: 700;
        color: #F8FAFC;
        margin: 6px 0;
    }
    .metric-sub {
        font-size: 0.8rem;
        font-weight: 500;
    }
    
    /* Badges */
    .badge-vip {
        background-color: rgba(16, 185, 129, 0.15);
        color: #10B981;
        border: 1px solid #10B981;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 700;
        font-size: 0.85rem;
    }
    .badge-potential {
        background-color: rgba(56, 189, 248, 0.15);
        color: #38BDF8;
        border: 1px solid #38BDF8;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 700;
        font-size: 0.85rem;
    }
    .badge-junk {
        background-color: rgba(239, 68, 68, 0.15);
        color: #EF4444;
        border: 1px solid #EF4444;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 700;
        font-size: 0.85rem;
    }
    
    /* Workflow Steps Card */
    .wf-step {
        background: #1E293B;
        border: 1px solid #334155;
        border-radius: 12px;
        padding: 16px;
        text-align: center;
    }
    .wf-step-num {
        background: #38BDF8;
        color: #0F172A;
        font-weight: 800;
        width: 32px;
        height: 32px;
        border-radius: 50%;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        margin-bottom: 8px;
    }
    
    /* Tab Styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #1E293B;
        border-radius: 8px;
        padding: 8px 20px;
        color: #94A3B8;
        border: 1px solid #334155;
    }
    .stTabs [aria-selected="true"] {
        background-color: #0284C7 !important;
        color: #FFFFFF !important;
        border-color: #38BDF8 !important;
    }
</style>
""", unsafe_allow_html=True)

DEFAULT_SHEET_ID = "1zLWzZT3a0qLL-Km66DVamoqBnduHvwJ0lUUZ3tD1qL0"

@st.cache_data(ttl=300)
def load_leads_from_google_sheet(sheet_id):
    """Tải dữ liệu live từ Google Sheets qua CSV gviz endpoint."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv"
    try:
        response = requests.get(url, timeout=12)
        if response.status_code == 200 and len(response.text) > 30 and not "html" in response.headers.get("content-type", ""):
            df = pd.read_csv(io.StringIO(response.text), encoding='utf-8-sig')
            df = df.dropna(how='all', axis=1)
            return df, True, "Thành công (Kết nối Google Sheets Trực tiếp)"
    except Exception as e:
        pass
    
    # Fallback to local test.csv if offline
    if os.path.exists("test.csv"):
        df = pd.read_csv("test.csv", encoding='utf-8-sig')
        df = df.dropna(how='all', axis=1)
        return df, False, "Sử dụng bản sao Dữ liệu Local (Offline)"
        
    return pd.DataFrame(), False, "Không thể tải dữ liệu"

# --- SIDEBAR ---
with st.sidebar:
    st.image("https://img.icons8.com/color/96/real-estate.png", width=64)
    st.title("AI Lead Scoring System")
    st.caption("Ngành Bất Động Sản • MindX Technology")
    st.markdown("---")
    
    st.subheader("📡 Nguồn Dữ Liệu")
    sheet_id_input = st.text_input("Google Sheet ID", value=DEFAULT_SHEET_ID)
    
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        if st.button("🔄 Đồng Bộ", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
            
    uploaded_file = st.file_uploader("Hoặc Tải File CSV/Excel", type=["csv", "xlsx"])
    
    st.markdown("---")
    st.subheader("⚙️ Cấu Hình Chấm Điểm AI")
    vip_threshold = st.slider("Ngưỡng VIP / Siêu Tiềm Năng (Điểm)", min_value=70, max_value=100, value=80, step=5)
    junk_threshold = st.slider("Ngưỡng Rác / Bỏ Qua (Điểm)", min_value=10, max_value=50, value=40, step=5)
    
    st.markdown("---")
    st.caption("📌 **Hệ thống AI Lead Scoring v2.0**")
    st.caption("Phát triển bởi AI Agent - MindX Real Estate Automation")

# --- DATA PROCESSING ---
if uploaded_file is not None:
    try:
        if uploaded_file.name.endswith(".csv"):
            df_raw = pd.read_csv(uploaded_file, encoding='utf-8-sig')
        else:
            df_raw = pd.read_excel(uploaded_file)
        data_source_msg = f"Đã tải file tải lên: {uploaded_file.name}"
        is_live = False
    except Exception as e:
        st.error(f"Lỗi đọc file: {e}")
        df_raw, is_live, data_source_msg = load_leads_from_google_sheet(sheet_id_input)
else:
    df_raw, is_live, data_source_msg = load_leads_from_google_sheet(sheet_id_input)

if df_raw.empty:
    st.error("⚠️ Không thể tải dữ liệu khách hàng. Vui lòng kiểm tra lại link Google Sheets hoặc file tải lên!")
    st.stop()

# Ensure standard columns
required_cols = ['id', 'ten_khach', 'sdt', 'nhu_cau_mo_ta']
for col in required_cols:
    if col not in df_raw.columns:
        # Try best effort column matching
        if col == 'id' and 'ID' in df_raw.columns:
            df_raw['id'] = df_raw['ID']
        elif col == 'ten_khach' and 'Tên' in df_raw.columns:
            df_raw['ten_khach'] = df_raw['Tên']
        elif col == 'sdt' and 'SĐT' in df_raw.columns:
            df_raw['sdt'] = df_raw['SĐT']
        elif col == 'nhu_cau_mo_ta' and 'Mô tả' in df_raw.columns:
            df_raw['nhu_cau_mo_ta'] = df_raw['Mô tả']

# Initialize session state for Human-in-the-loop edits
if 'leads_df' not in st.session_state or st.session_state.get('data_hash') != hash(str(df_raw.values.tobytes())):
    # Run AI Scoring on all rows
    scored_rows = []
    for idx, row in df_raw.iterrows():
        desc = str(row.get('nhu_cau_mo_ta', ''))
        score_res = analyze_and_score_lead(desc)
        
        # Override tier based on custom thresholds
        ai_score = score_res['ai_score']
        if ai_score >= vip_threshold:
            ai_tier = "VIP / Siêu tiềm năng"
        elif ai_score >= junk_threshold:
            ai_tier = "Tiềm năng"
        else:
            ai_tier = "Không tiềm năng / Rác"

        # Default Human Status
        if ai_tier == "VIP / Siêu tiềm năng":
            default_human_status = "Đã duyệt (Sẵn sàng bàn giao)"
        elif ai_tier == "Không tiềm năng / Rác":
            default_human_status = "Bỏ qua / Rác"
        else:
            default_human_status = "Chờ duyệt"

        scored_rows.append({
            'id': str(row.get('id', idx + 1)),
            'ten_khach': str(row.get('ten_khach', 'Khách vô danh')),
            'sdt': str(row.get('sdt', '')),
            'nhu_cau_mo_ta': desc,
            'ai_score': int(ai_score),
            'ai_tier': ai_tier,
            'ai_summary': score_res['ai_summary'],
            'plus_reasons': ", ".join(score_res['plus_reasons']),
            'minus_reasons': ", ".join(score_res['minus_reasons']),
            'matched_keywords': ", ".join(score_res['matched_keywords']),
            'human_score': int(ai_score),
            'human_status': default_human_status,
            'human_note': ""
        })

    st.session_state.leads_df = pd.DataFrame(scored_rows)
    st.session_state.data_hash = hash(str(df_raw.values.tobytes()))

df_leads = st.session_state.leads_df

# --- HEADER APP ---
st.markdown(f"""
<div class="header-box">
    <div style="display: flex; justify-content: space-between; align-items: center;">
        <div>
            <h1 style="margin:0; font-size: 2rem; color: #F8FAFC;">🏠 Hệ Thống AI Lead Scoring & Automation BĐS</h1>
            <p style="margin: 5px 0 0 0; color: #94A3B8;">Tự động hóa phân tích nhu cầu, chấm điểm khách hàng tiềm năng & Kiểm duyệt Human-in-the-loop</p>
        </div>
        <div>
            <span style="background: #1E293B; border: 1px solid #334155; padding: 8px 16px; border-radius: 20px; font-size: 0.85rem; color: #38BDF8; font-weight: 600;">
                📡 {data_source_msg}
            </span>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# --- TOP SUMMARY METRICS ---
c1, c2, c3, c4, c5 = st.columns(5)

total_leads = len(df_leads)
vip_leads = len(df_leads[df_leads['ai_tier'] == 'VIP / Siêu tiềm năng'])
potential_leads = len(df_leads[df_leads['ai_tier'] == 'Tiềm năng'])
junk_leads = len(df_leads[df_leads['ai_tier'] == 'Không tiềm năng / Rác'])
approved_leads = len(df_leads[df_leads['human_status'] == 'Đã duyệt (Sẵn sàng bàn giao)'])

with c1:
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-title">Tổng Lead Nhận</div>
        <div class="metric-value">{total_leads}</div>
        <div class="metric-sub" style="color: #94A3B8;">Dữ liệu Google Sheets</div>
    </div>
    """, unsafe_allow_html=True)

with c2:
    st.markdown(f"""
    <div class="metric-card" style="border-left: 4px solid #10B981;">
        <div class="metric-title">VIP / Siêu Tiềm Năng</div>
        <div class="metric-value" style="color: #10B981;">{vip_leads}</div>
        <div class="metric-sub" style="color: #10B981;">▲ {(vip_leads/total_leads*100 if total_leads else 0):.1f}% tổng số lead</div>
    </div>
    """, unsafe_allow_html=True)

with c3:
    st.markdown(f"""
    <div class="metric-card" style="border-left: 4px solid #38BDF8;">
        <div class="metric-title">Khách Tiềm Năng</div>
        <div class="metric-value" style="color: #38BDF8;">{potential_leads}</div>
        <div class="metric-sub" style="color: #38BDF8;">● Phân khúc tầm trung</div>
    </div>
    """, unsafe_allow_html=True)

with c4:
    st.markdown(f"""
    <div class="metric-card" style="border-left: 4px solid #EF4444;">
        <div class="metric-title">Không Tiềm Năng / Rác</div>
        <div class="metric-value" style="color: #EF4444;">{junk_leads}</div>
        <div class="metric-sub" style="color: #EF4444;">▼ Đã lọc bỏ tự động</div>
    </div>
    """, unsafe_allow_html=True)

with c5:
    st.markdown(f"""
    <div class="metric-card" style="border-left: 4px solid #F59E0B;">
        <div class="metric-title">Đã Chốt & Bàn Giao</div>
        <div class="metric-value" style="color: #F59E0B;">{approved_leads}</div>
        <div class="metric-sub" style="color: #F59E0B;">Sẵn sàng xuất Excel</div>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# --- MAIN TABS ---
tab_wf, tab_hitl, tab_analytics, tab_export = st.tabs([
    "🚀 Quy Trình Tổng Quan",
    "👥 Human-in-the-Loop (Kiểm Duyệt Lead)",
    "📊 Thống Kê & Phân Tích",
    "📥 Chốt Kết Quả & Xuất Excel"
])

# ---------------------------------------------------------
# TAB 1: WORKFLOW OVERVIEW
# ---------------------------------------------------------
with tab_wf:
    st.markdown("### 🔄 Quy Trình Tổng Quan Hệ Thống AI Lead Scoring & Automation")
    st.caption("Quy trình 5 bước khép kín kết hợp Trí Tuệ Nhân Tạo & Kiểm Duyệt Con Người (Human-in-the-Loop)")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    w1, w2, w3, w4, w5 = st.columns(5)
    
    with w1:
        st.markdown("""
        <div class="wf-step">
            <div class="wf-step-num">1</div>
            <h4 style="margin: 8px 0; color: #38BDF8;">LẤY DỮ LIỆU</h4>
            <p style="font-size: 0.8rem; color: #94A3B8;">Tự động kết nối Google Sheets API / CSV Live Sync</p>
            <span style="font-size: 0.75rem; background: #0F172A; padding: 4px 8px; border-radius: 4px; color: #10B981;">● Tự động 100%</span>
        </div>
        """, unsafe_allow_html=True)
        
    with w2:
        st.markdown("""
        <div class="wf-step">
            <div class="wf-step-num">2</div>
            <h4 style="margin: 8px 0; color: #38BDF8;">AI LEAD SCORING</h4>
            <p style="font-size: 0.8rem; color: #94A3B8;">Phân tích ngữ cảnh, từ khóa & Quy tắc nghiệp vụ BĐS (+50 / -50)</p>
            <span style="font-size: 0.75rem; background: #0F172A; padding: 4px 8px; border-radius: 4px; color: #10B981;">● AI NLP Engine</span>
        </div>
        """, unsafe_allow_html=True)

    with w3:
        st.markdown("""
        <div class="wf-step">
            <div class="wf-step-num">3</div>
            <h4 style="margin: 8px 0; color: #F59E0B;">HUMAN IN THE LOOP</h4>
            <p style="font-size: 0.8rem; color: #94A3B8;">Chăm sóc viên kiểm duyệt, điều chỉnh điểm & ghi chú trạng thái</p>
            <span style="font-size: 0.75rem; background: #0F172A; padding: 4px 8px; border-radius: 4px; color: #F59E0B;">👤 Sale Lead Review</span>
        </div>
        """, unsafe_allow_html=True)

    with w4:
        st.markdown("""
        <div class="wf-step">
            <div class="wf-step-num">4</div>
            <h4 style="margin: 8px 0; color: #10B981;">CHỐT KẾT QUẢ</h4>
            <p style="font-size: 0.8rem; color: #94A3B8;">Xác nhận danh sách lead tối ưu chuẩn bị phân bổ Sale</p>
            <span style="font-size: 0.75rem; background: #0F172A; padding: 4px 8px; border-radius: 4px; color: #10B981;">✔ Final Approval</span>
        </div>
        """, unsafe_allow_html=True)

    with w5:
        st.markdown("""
        <div class="wf-step">
            <div class="wf-step-num">5</div>
            <h4 style="margin: 8px 0; color: #38BDF8;">XUẤT EXCEL</h4>
            <p style="font-size: 0.8rem; color: #94A3B8;">Xuất dữ liệu định dạng chuẩn sẵn sàng bàn giao</p>
            <span style="font-size: 0.75rem; background: #0F172A; padding: 4px 8px; border-radius: 4px; color: #38BDF8;">📥 Export .XLSX</span>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br><hr><br>", unsafe_allow_html=True)

    st.subheader("📜 Tiêu Chí Chấm Điểm AI Đã Được Thiết Lập")
    
    col_rule1, col_rule2, col_rule3 = st.columns(3)
    
    with col_rule1:
        st.success("🟢 **TIÊU CHÍ CỘNG 50 ĐIỂM (VIP)**")
        st.markdown("""
        - **Ngân sách lớn**: Số tiền $\ge$ 20 tỷ hoặc từ khóa *"tài chính mạnh"*, *"không thành vấn đề"*.
        - **Loại hình cao cấp**: Biệt thự đơn lập, Penthouse, Shophouse mặt đường lớn, Quỹ đất công nghiệp, Sàn văn phòng lớn.
        - **Vị trí đắc địa**: Quận 1, Ven sông, Vinhomes Ocean Park, Phú Mỹ Hưng.
        - **Đối tượng KH**: Chủ doanh nghiệp, Nhà đầu tư chuyên nghiệp, Mua sỉ, Mua số lượng lớn.
        - **Minh bạch & Cấp thiết**: Pháp lý chuẩn 100%, Sổ hồng riêng, Muốn gặp trực tiếp chủ đầu tư.
        """)

    with col_rule2:
        st.error("🔴 **TIÊU CHÍ TRỪ 50 ĐIỂM (RÁC)**")
        st.markdown("""
        - **Yêu cầu phi thực tế**: BĐS giá thấp vô lý (VD: Nhà Q1 giá 1-2 tỷ, nhà trung tâm hồ bơi vài trăm triệu, thuê nguyên căn 2 triệu...).
        - **Không có nhu cầu**: Nhầm số, Không có nhu cầu, Dữ liệu cũ, Nhầm ngành.
        - **Không thiện chí**: Hỏi giá cho vui, Chưa có ý định mua, Thái độ không hợp tác.
        - **Spam / Quảng cáo**: Bảo hiểm, Vay vốn, Mời chào dịch vụ.
        - **Lỗi liên lạc**: Thuê bao, Gọi nhiều lần không nghe máy, Không phản hồi Zalo.
        """)

    with col_rule3:
        st.info("🔵 **TRƯỜNG HỢP KHÁC (TẦM TRUNG)**")
        st.markdown("""
        - Khách hàng tìm mua **chung cư, nhà phố tầm trung (3-10 tỷ)**.
        - Khách hàng **cần vay ngân hàng**, đang cân nhắc chính sách chiết khấu.
        - Khách hàng có **nhu cầu thực** nhưng cần tư vấn thêm về pháp lý hoặc vị trí.
        - *Cộng 10 điểm thưởng cho nhu cầu thực tế phân khúc tầm trung.*
        """)

# ---------------------------------------------------------
# TAB 2: HUMAN-IN-THE-LOOP (KIỂM DUYỆT LEAD)
# ---------------------------------------------------------
with tab_hitl:
    st.markdown("### 👥 Human-in-the-Loop: Kiểm Duyệt & Điều Chỉnh Điểm Lead")
    st.caption("Cho phép Chăm sóc viên / Sale Lead kiểm tra giải trình AI, sửa điểm số và chốt trạng thái phân bổ.")
    
    # Filter Controls
    f_col1, f_col2, f_col3, f_col4 = st.columns([2, 2, 2, 3])
    
    with f_col1:
        filter_tier = st.selectbox("Phân Loại AI", ["Tất cả Tầng Lead", "VIP / Siêu tiềm năng", "Tiềm năng", "Không tiềm năng / Rác"])
    with f_col2:
        filter_status = st.selectbox("Trạng Thái Kiểm Duyệt", ["Tất cả Trạng Thái", "Chờ duyệt", "Đã duyệt (Sẵn sàng bàn giao)", "Cần gọi lại gấp", "Bỏ qua / Rác"])
    with f_col3:
        search_query = st.text_input("🔍 Tìm Tên / SĐT / Từ khóa", value="")
    with f_col4:
        st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
        b_col1, b_col2 = st.columns(2)
        with b_col1:
            if st.button("⚡ Duyệt Nhanh VIP", use_container_width=True):
                st.session_state.leads_df.loc[st.session_state.leads_df['ai_tier'] == 'VIP / Siêu tiềm năng', 'human_status'] = 'Đã duyệt (Sẵn sàng bàn giao)'
                st.success("Đã duyệt toàn bộ Lead VIP!")
                st.rerun()
        with b_col2:
            if st.button("🔄 Reset Trạng Thái", use_container_width=True):
                for idx, row in st.session_state.leads_df.iterrows():
                    tier = row['ai_tier']
                    if tier == "VIP / Siêu tiềm năng":
                        st.session_state.leads_df.at[idx, 'human_status'] = "Đã duyệt (Sẵn sàng bàn giao)"
                    elif tier == "Không tiềm năng / Rác":
                        st.session_state.leads_df.at[idx, 'human_status'] = "Bỏ qua / Rác"
                    else:
                        st.session_state.leads_df.at[idx, 'human_status'] = "Chờ duyệt"
                st.info("Đã reset trạng thái duyệt về mặc định AI!")
                st.rerun()

    # Apply Filters
    df_filtered = df_leads.copy()
    
    if filter_tier != "Tất cả Tầng Lead":
        df_filtered = df_filtered[df_filtered['ai_tier'] == filter_tier]
    if filter_status != "Tất cả Trạng Thái":
        df_filtered = df_filtered[df_filtered['human_status'] == filter_status]
    if search_query.strip():
        q = search_query.strip().lower()
        df_filtered = df_filtered[
            df_filtered['ten_khach'].str.lower().str.contains(q) |
            df_filtered['sdt'].astype(str).str.contains(q) |
            df_filtered['nhu_cau_mo_ta'].str.lower().str.contains(q) |
            df_filtered['matched_keywords'].str.lower().str.contains(q)
        ]

    st.markdown(f"**Hiển thị {len(df_filtered)} / {len(df_leads)} khách hàng**")
    
    view_mode = st.radio("Chế độ hiển thị", ["📝 Bảng Dữ Liệu Tương Tác (Bulk Edit)", "🎴 Thẻ Chi Tiết Từng Lead"], horizontal=True)

    if view_mode == "📝 Bảng Dữ Liệu Tương Tác (Bulk Edit)":
        st.caption("💡 Mẹo: Bạn có thể chỉnh sửa trực tiếp cột 'human_score', 'human_status' và 'human_note' ngay trong bảng dưới đây!")
        
        # Configure columns for st.data_editor
        edited_df = st.data_editor(
            df_filtered[['id', 'ten_khach', 'sdt', 'ai_score', 'ai_tier', 'human_score', 'human_status', 'human_note', 'ai_summary', 'nhu_cau_mo_ta']],
            column_config={
                "id": st.column_config.TextColumn("ID", disabled=True, width="small"),
                "ten_khach": st.column_config.TextColumn("Tên Khách", disabled=True, width="medium"),
                "sdt": st.column_config.TextColumn("SĐT", disabled=True, width="medium"),
                "ai_score": st.column_config.NumberColumn("Điểm AI", disabled=True, format="%d điểm"),
                "ai_tier": st.column_config.TextColumn("Tầng AI", disabled=True),
                "human_score": st.column_config.NumberColumn("Điểm Duyệt", min_value=0, max_value=100, step=5, format="%d điểm"),
                "human_status": st.column_config.SelectboxColumn(
                    "Trạng Thái Duyệt",
                    options=["Chờ duyệt", "Đã duyệt (Sẵn sàng bàn giao)", "Cần gọi lại gấp", "Bỏ qua / Rác"],
                    required=True
                ),
                "human_note": st.column_config.TextColumn("Ghi Chú Nhân Sự", width="large"),
                "ai_summary": st.column_config.TextColumn("Giải Trình AI", disabled=True, width="large"),
                "nhu_cau_mo_ta": st.column_config.TextColumn("Mô Tả Nhu Cầu", disabled=True, width="large")
            },
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            key="bulk_data_editor"
        )
        
        # Save edits back to session state dataframe
        if not edited_df.equals(df_filtered[['id', 'ten_khach', 'sdt', 'ai_score', 'ai_tier', 'human_score', 'human_status', 'human_note', 'ai_summary', 'nhu_cau_mo_ta']]):
            for idx, edited_row in edited_df.iterrows():
                lead_id = edited_row['id']
                match_idx = st.session_state.leads_df[st.session_state.leads_df['id'] == lead_id].index
                if len(match_idx) > 0:
                    real_idx = match_idx[0]
                    st.session_state.leads_df.at[real_idx, 'human_score'] = edited_row['human_score']
                    st.session_state.leads_df.at[real_idx, 'human_status'] = edited_row['human_status']
                    st.session_state.leads_df.at[real_idx, 'human_note'] = str(edited_row['human_note'] if pd.notna(edited_row['human_note']) else "")
            st.toast("Đã lưu các thay đổi!", icon="✅")

    else:
        # Card Detailed View
        st.markdown("---")
        for idx, row in df_filtered.iterrows():
            lead_id = row['id']
            tier = row['ai_tier']
            score = row['ai_score']
            
            if tier == "VIP / Siêu tiềm năng":
                badge_html = f'<span class="badge-vip">🌟 VIP ({score} Điểm)</span>'
                border_color = "#10B981"
            elif tier == "Tiềm năng":
                badge_html = f'<span class="badge-potential">✅ TIỀM NĂNG ({score} Điểm)</span>'
                border_color = "#38BDF8"
            else:
                badge_html = f'<span class="badge-junk">⛔ RÁC ({score} Điểm)</span>'
                border_color = "#EF4444"

            with st.container():
                st.markdown(f"""
                <div style="background: #1E293B; border-left: 6px solid {border_color}; border-radius: 10px; padding: 16px; margin-bottom: 16px;">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <h3 style="margin: 0; color: #F8FAFC;">#{lead_id} • {row['ten_khach']} - 📞 {row['sdt']}</h3>
                        <div>{badge_html}</div>
                    </div>
                    <p style="margin: 10px 0; color: #CBD5E1; background: #0F172A; padding: 12px; border-radius: 8px; font-style: italic;">
                        "{row['nhu_cau_mo_ta']}"
                    </p>
                    <p style="margin: 5px 0; color: #94A3B8; font-size: 0.9rem;">
                        <strong>🤖 AI Đánh giá:</strong> {row['ai_summary']}
                    </p>
                </div>
                """, unsafe_allow_html=True)

                col_edit1, col_edit2, col_edit3 = st.columns([2, 3, 4])
                with col_edit1:
                    new_score = st.number_input(f"Điểm Chốt (Lead #{lead_id})", min_value=0, max_value=100, value=int(row['human_score']), key=f"score_{lead_id}")
                with col_edit2:
                    new_status = st.selectbox(
                        f"Trạng Thái (Lead #{lead_id})",
                        options=["Chờ duyệt", "Đã duyệt (Sẵn sàng bàn giao)", "Cần gọi lại gấp", "Bỏ qua / Rác"],
                        index=["Chờ duyệt", "Đã duyệt (Sẵn sàng bàn giao)", "Cần gọi lại gấp", "Bỏ qua / Rác"].index(row['human_status']) if row['human_status'] in ["Chờ duyệt", "Đã duyệt (Sẵn sàng bàn giao)", "Cần gọi lại gấp", "Bỏ qua / Rác"] else 0,
                        key=f"status_{lead_id}"
                    )
                with col_edit3:
                    new_note = st.text_input(f"Ghi chú Sales (Lead #{lead_id})", value=str(row['human_note']), key=f"note_{lead_id}")

                # Sync card edits to dataframe
                real_idx = st.session_state.leads_df[st.session_state.leads_df['id'] == lead_id].index[0]
                st.session_state.leads_df.at[real_idx, 'human_score'] = new_score
                st.session_state.leads_df.at[real_idx, 'human_status'] = new_status
                st.session_state.leads_df.at[real_idx, 'human_note'] = new_note
                st.markdown("<br>", unsafe_allow_html=True)

# ---------------------------------------------------------
# TAB 3: ANALYTICS & THỐNG KÊ
# ---------------------------------------------------------
with tab_analytics:
    st.markdown("### 📊 Báo Cáo Phân Tích Lead & Phân Bổ Chất Lượng")
    st.caption("Thống kê chi tiết chất lượng tệp dữ liệu khách hàng từ Google Sheets")
    
    col_chart1, col_chart2 = st.columns(2)
    
    with col_chart1:
        st.subheader("🍩 Phân Bổ Tầng Chất Lượng Lead (AI Tier)")
        tier_counts = df_leads['ai_tier'].value_counts().reset_index()
        tier_counts.columns = ['ai_tier', 'count']
        
        fig_donut = px.pie(
            tier_counts, 
            names='ai_tier', 
            values='count', 
            hole=0.5,
            color='ai_tier',
            color_discrete_map={
                'VIP / Siêu tiềm năng': '#10B981',
                'Tiềm năng': '#38BDF8',
                'Không tiềm năng / Rác': '#EF4444'
            }
        )
        fig_donut.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(color='#E2E8F0'))
        st.plotly_chart(fig_donut, use_container_width=True)

    with col_chart2:
        st.subheader("📈 Biểu Đồ Phân Bổ Điểm Số (Score Distribution)")
        fig_hist = px.histogram(
            df_leads,
            x='ai_score',
            nbins=10,
            color='ai_tier',
            color_discrete_map={
                'VIP / Siêu tiềm năng': '#10B981',
                'Tiềm năng': '#38BDF8',
                'Không tiềm năng / Rác': '#EF4444'
            },
            labels={'ai_score': 'Điểm số AI', 'count': 'Số lượng Lead'}
        )
        fig_hist.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(color='#E2E8F0'))
        st.plotly_chart(fig_hist, use_container_width=True)

    st.markdown("---")
    col_chart3, col_chart4 = st.columns(2)
    
    with col_chart3:
        st.subheader("🌟 Top Từ Khóa VIP Được Nhận Diện Nổi Bật")
        # Extract matched keywords for VIPs
        vip_keywords = []
        for kw_str in df_leads[df_leads['ai_tier'] == 'VIP / Siêu tiềm năng']['matched_keywords']:
            if kw_str:
                vip_keywords.extend([k.strip() for k in kw_str.split(",") if k.strip()])
        
        if vip_keywords:
            df_kw = pd.Series(vip_keywords).value_counts().head(8).reset_index()
            df_kw.columns = ['keyword', 'frequency']
            fig_bar_kw = px.bar(df_kw, x='frequency', y='keyword', orientation='h', color_discrete_sequence=['#10B981'])
            fig_bar_kw.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(color='#E2E8F0'), yaxis=dict(autorange="reversed"))
            st.plotly_chart(fig_bar_kw, use_container_width=True)
        else:
            st.info("Chưa có dữ liệu từ khóa VIP")

    with col_chart4:
        st.subheader("📌 Phân Bổ Trạng Thái Kiểm Duyệt Human-in-the-Loop")
        status_counts = df_leads['human_status'].value_counts().reset_index()
        status_counts.columns = ['human_status', 'count']
        
        fig_status = px.bar(
            status_counts, 
            x='count', 
            y='human_status', 
            orientation='h', 
            color='human_status',
            color_discrete_map={
                'Đã duyệt (Sẵn sàng bàn giao)': '#10B981',
                'Chờ duyệt': '#F59E0B',
                'Cần gọi lại gấp': '#38BDF8',
                'Bỏ qua / Rác': '#EF4444'
            }
        )
        fig_status.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font=dict(color='#E2E8F0'), yaxis=dict(autorange="reversed"))
        st.plotly_chart(fig_status, use_container_width=True)

# ---------------------------------------------------------
# TAB 4: EXPORT EXCEL & CHỐT KẾT QUẢ
# ---------------------------------------------------------
with tab_export:
    st.markdown("### 📥 Chốt Kết Quả & Xuất File Excel Bàn Giao")
    st.caption("Xuất toàn bộ danh sách khách hàng đã qua xử lý AI và Human-in-the-loop ra file Excel định dạng chuẩn.")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Filter final decision options
    export_option = st.radio(
        "Lựa chọn tập dữ liệu xuất file Excel:",
        ["✅ Tất cả Lead Đã Duyệt (Sẵn sàng bàn giao)", "📋 Toàn bộ Danh sách Lead (Bao gồm cả VIP, Tiềm năng & Rác)"],
        index=0
    )
    
    if "Tất cả Lead Đã Duyệt" in export_option:
        df_export = df_leads[df_leads['human_status'] == 'Đã duyệt (Sẵn sàng bàn giao)'].copy()
    else:
        df_export = df_leads.copy()

    st.info(f"📊 Tập dữ liệu xuất sẽ bao gồm **{len(df_export)} khách hàng**.")
    
    # Preview Table
    st.markdown("#### 👁️ Xem Trước Dữ Liệu Bàn Giao Excel")
    st.dataframe(
        df_export[['id', 'ten_khach', 'sdt', 'ai_score', 'ai_tier', 'human_score', 'human_status', 'human_note', 'nhu_cau_mo_ta']],
        use_container_width=True
    )
    
    # Function to generate styled Excel workbook
    def generate_excel_bytes(df_to_export):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bàn Giao Lead BĐS"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Header formatting
        headers = [
            "STT / ID", "Tên Khách Hàng", "Số Điện Thoại", "Mô Tả Nhu Cầu",
            "Điểm AI", "Phân Loại AI", "Giải Trình AI",
            "Điểm Chốt (Human)", "Trạng Thái Duyệt", "Ghi Chú Sales"
        ]
        
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        ws.row_dimensions[1].height = 28
        
        # Fills for tiers
        vip_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid") # Soft green
        pot_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid") # Soft blue
        junk_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid") # Soft red
        
        vip_font = Font(name="Calibri", size=11, bold=True, color="137333")
        pot_font = Font(name="Calibri", size=11, bold=True, color="1A73E8")
        junk_font = Font(name="Calibri", size=11, bold=True, color="C5221F")

        # Append Rows
        for r_idx, row in df_to_export.iterrows():
            row_data = [
                row['id'],
                row['ten_khach'],
                str(row['sdt']),
                row['nhu_cau_mo_ta'],
                row['ai_score'],
                row['ai_tier'],
                row['ai_summary'],
                row['human_score'],
                row['human_status'],
                row['human_note']
            ]
            ws.append(row_data)
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 24
            
            # Apply styling
            tier = row['ai_tier']
            for c_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=current_row, column=c_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                
                # Apply tier highlight to AI Tier column (column 6) and Score column (column 5)
                if c_idx in [5, 6]:
                    if tier == "VIP / Siêu tiềm năng":
                        cell.fill = vip_fill
                        cell.font = vip_font
                    elif tier == "Tiềm năng":
                        cell.fill = pot_fill
                        cell.font = pot_font
                    else:
                        cell.fill = junk_fill
                        cell.font = junk_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        # Set auto column widths
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Specific column width overrides
        ws.column_dimensions['D'].width = 45 # Description
        ws.column_dimensions['G'].width = 50 # AI Summary
        ws.column_dimensions['J'].width = 30 # Note

        output_stream = io.BytesIO()
        wb.save(output_stream)
        return output_stream.getvalue()

    excel_bytes = generate_excel_bytes(df_export)
    
    st.markdown("<br>", unsafe_allow_html=True)
    st.download_button(
        label="📥 Tải File Excel Bàn Giao (.xlsx)",
        data=excel_bytes,
        file_name="Danh_Sach_Khach_Hang_Da_Cham_Diem_Lead_Scoring.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
